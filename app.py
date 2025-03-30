from flask import Flask, request, send_file, render_template_string
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib import rcParams

rcParams["font.family"] = "Times New Roman"

app = Flask(__name__)

UPLOAD_FOLDER = "uploads"
RESULT_FOLDER = "results"
SECOND_UPLOAD_FOLDER = "uploads_second"
SECOND_RESULT_FOLDER = "results_second"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULT_FOLDER, exist_ok=True)
os.makedirs(SECOND_UPLOAD_FOLDER, exist_ok=True)
os.makedirs(SECOND_RESULT_FOLDER, exist_ok=True)

HTML_TEMPLATE = '''
<!doctype html>
<html>
<head><title>KPT BC Liên Chiểu</title></head>
<body>
    <h2>KPI PTC lũy kế Bưu Cục Liên Chiểu</h2>
    <form action="/process" method="post" enctype="multipart/form-data">
        <input type="file" name="file"><br><br>
        <label>Tỷ lệ COD tối thiểu:</label>
        <input type="number" name="cod_threshold" step="0.1" value="95"><br><br>
        <label>Tỷ lệ đúng giờ tối thiểu:</label>
        <input type="number" name="time_threshold" step="0.1" value="85"><br><br>
        <input type="submit" value="Upload and Process">
    </form>
    <br>
    {% if img_path %}
        <h2>Processed Cảnh báo KPI PTC lũy kế Bưu Cục Liên Chiểu</h2>
        <img src="{{ img_path }}" style="max-width: 30%;"><br><br>
        <form action="/download" method="get">
            <button type="submit">Xuất file Excel</button>
        </form>
    {% endif %}
    <br><br><a href="/second">Chuyển sang chương trình xử lý thứ hai</a>
</body>
</html>
'''

SECOND_TEMPLATE = '''
<!doctype html>
<html>
<head><title>KPI đúng giờ Bưu cục Liên Chiểut</title></head>
<body>
    <h2>KPI đúng giờ Bưu cục Liên Chiểu</h2>
    <form action="/second/process" method="post" enctype="multipart/form-data">
        <input type="file" name="file"><br><br>
        <input type="submit" value="Upload và xử lý">
    </form>
    <br>
    {% if img_path %}
        <h2>Cảnh báo KPI đúng giờ Bưu cục Liên Chiểu</h2>
        <img src="{{ img_path }}" style="max-width: 80%;"><br><br>
        <form action="/second/download" method="get">
            <button type="submit">Tải file Excel kết quả</button>
        </form>
    {% endif %}
    <br><br><a href="/">Quay lại chương trình đầu</a>
</body>
</html>
'''

@app.route('/process', methods=['POST'])
def process_file():
    if 'file' not in request.files:
        return "No file part"
    file = request.files['file']
    if file.filename == '':
        return "No selected file"

    cod_threshold = float(request.form.get('cod_threshold', 95))
    time_threshold = float(request.form.get('time_threshold', 85))

    file_path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(file_path)

    # Lấy ngưỡng từ form người dùng (mặc định 85 nếu không nhập)
    try:
        ngay_threshold = float(request.form.get('ngay_threshold', 85))
        thang_threshold = float(request.form.get('thang_threshold', 85))
    except:
        ngay_threshold = 85
        thang_threshold = 85

    # Lấy ngưỡng từ form người dùng
    try:
        ngay_threshold = float(request.form.get('ngay_threshold', 85))
        thang_threshold = float(request.form.get('thang_threshold', 85))
    except:
        ngay_threshold = 85
        thang_threshold = 85

    # Lấy ngưỡng từ form
    try:
        ngay_threshold = float(request.form.get('ngay_threshold', 85))
        thang_threshold = float(request.form.get('thang_threshold', 85))
    except:
        ngay_threshold = 85
        thang_threshold = 85

    # Lấy ngưỡng từ form người dùng
    ngay_threshold = float(request.form.get('ngay_threshold', 85))
    thang_threshold = float(request.form.get('thang_threshold', 85))

    # Gán ngưỡng cố định (bỏ nhập từ người dùng)
    ngay_threshold = 85
    thang_threshold = 85

    # Lấy ngưỡng từ form người dùng
    
    

    # Lấy ngưỡng từ form (mặc định là 85 nếu không nhập)
    ngay_threshold = float(request.form.get('ngay_threshold', 85))
    thang_threshold = float(request.form.get('thang_threshold', 85))

    # Lấy ngưỡng từ form
    ngay_threshold = float(request.form.get('ngay_threshold', 85))
    thang_threshold = float(request.form.get('thang_threshold', 85))

    # Lấy ngưỡng từ form (mặc định 85 nếu không nhập)
    ngay_threshold = float(request.form.get('ngay_threshold', 85))
    thang_threshold = float(request.form.get('thang_threshold', 85))

    # Lấy ngưỡng từ form người dùng
    ngay_threshold = float(request.form.get('ngay_threshold', 85))
    thang_threshold = float(request.form.get('thang_threshold', 85))

    # Lấy ngưỡng từ form (mặc định 85 nếu không nhập)
    ngay_threshold = float(request.form.get('ngay_threshold', 85))
    thang_threshold = float(request.form.get('thang_threshold', 85))

    # Lấy ngưỡng từ form người dùng (nếu không nhập thì mặc định là 85)
    ngay_threshold = float(request.form.get('ngay_threshold', 85))
    thang_threshold = float(request.form.get('thang_threshold', 85))

    # Lấy ngưỡng từ form (mặc định 85 nếu không nhập)
    ngay_threshold = float(request.form.get('ngay_threshold', 85))
    thang_threshold = float(request.form.get('thang_threshold', 85))

    # Lấy ngưỡng từ form (nếu không có thì dùng mặc định 85)
    ngay_threshold = float(request.form.get('ngay_threshold', 85))
    thang_threshold = float(request.form.get('thang_threshold', 85))

    # Lấy hoặc giữ lại ngưỡng cảnh báo
    ngay_threshold = float(request.form.get('ngay_threshold', last_ngay_threshold if 'last_ngay_threshold' in globals() else 85))
    thang_threshold = float(request.form.get('thang_threshold', last_thang_threshold if 'last_thang_threshold' in globals() else 85))
        
    # Lấy hoặc giữ lại ngưỡng cảnh báo
    ngay_threshold = float(request.form.get('ngay_threshold', last_ngay_threshold if 'last_ngay_threshold' in globals() else 85))
    thang_threshold = float(request.form.get('thang_threshold', last_thang_threshold if 'last_thang_threshold' in globals() else 85))
    last_ngay_threshold = ngay_threshold
    last_thang_threshold = thang_threshold

    # Lấy hoặc giữ lại ngưỡng cũ
    ngay_threshold = float(request.form.get('ngay_threshold', last_ngay_threshold if 'last_ngay_threshold' in globals() else 85))
    thang_threshold = float(request.form.get('thang_threshold', last_thang_threshold if 'last_thang_threshold' in globals() else 85))
    last_ngay_threshold = ngay_threshold
    last_thang_threshold = thang_threshold

    # Lấy hoặc giữ lại ngưỡng cũ
    ngay_threshold = float(request.form.get('ngay_threshold', last_ngay_threshold if 'last_ngay_threshold' in globals() else 85))
    thang_threshold = float(request.form.get('thang_threshold', last_thang_threshold if 'last_thang_threshold' in globals() else 85))
    last_ngay_threshold = ngay_threshold
    last_thang_threshold = thang_threshold

    ngay_threshold = float(request.form.get('ngay_threshold', 85))
    thang_threshold = float(request.form.get('thang_threshold', 85))

    ngay_threshold = float(request.form.get('ngay_threshold', 85))
    thang_threshold = float(request.form.get('thang_threshold', 85))

    df = pd.read_excel(file_path)
    df.columns = df.columns.str.strip()

    expected_columns = ["Tuyến", "Phát thành công COD", "Phát thành công đúng giờ"]
    if not all(col in df.columns for col in expected_columns):
        return f"Thiếu cột: {df.columns.tolist()}"

    df_filtered = df[["Tuyến", "Phát thành công COD", "Phát thành công đúng giờ"]].copy()
    df_filtered.columns = ["Nhân viên", "Tỷ lệ COD", "Tỷ lệ đúng giờ"]
    df_filtered["Tỷ lệ COD"] = pd.to_numeric(df_filtered["Tỷ lệ COD"], errors="coerce")
    df_filtered["Tỷ lệ đúng giờ"] = pd.to_numeric(df_filtered["Tỷ lệ đúng giờ"], errors="coerce")
    df_filtered = df_filtered[(df_filtered["Tỷ lệ COD"] >= 70) & (df_filtered["Tỷ lệ COD"] < 100)]
    df_filtered = df_filtered.sort_values(by="Tỷ lệ COD", ascending=False)

    output_path = os.path.join(RESULT_FOLDER, "KPI_Canh_Bao.xlsx")
    df_filtered.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for row in range(2, ws.max_row + 1):
        if ws[f"B{row}"].value < cod_threshold:
            ws[f"B{row}"].fill = red_fill
        if ws[f"C{row}"].value < time_threshold:
            ws[f"C{row}"].fill = red_fill
    wb.save(output_path)

    img_path = os.path.join(RESULT_FOLDER, "KPI_Canh_Bao.png")
    fig, ax = plt.subplots(figsize=(4, len(df_filtered) * 0.2))
    ax.axis('off')
    df_filtered["Tỷ lệ COD"] = df_filtered["Tỷ lệ COD"].apply(lambda x: f"{x:.2f}%")
    df_filtered["Tỷ lệ đúng giờ"] = df_filtered["Tỷ lệ đúng giờ"].apply(lambda x: f"{x:.2f}%")
    table = ax.table(cellText=df_filtered.values.tolist(), colLabels=df_filtered.columns, cellLoc='center', loc='center')
    table.auto_set_font_size(False)
    table.set_fontsize(9)
    table.auto_set_column_width([0, 1, 2])

    for key, cell in table.get_celld().items():
        if key[0] == 0:
            cell.set_facecolor('#4C72B0')
            cell.set_text_props(color='white', weight='bold')
        elif key[0] > 0:
            value_cod = float(df_filtered.iloc[key[0]-1, 1].replace('%',''))
            value_time = float(df_filtered.iloc[key[0]-1, 2].replace('%',''))
            if (key[1] == 1 and value_cod < cod_threshold) or (key[1] == 2 and value_time < time_threshold):
                cell.set_facecolor("red")
                cell.set_text_props(color='white')
    plt.savefig(img_path, bbox_inches='tight', dpi=300)

    return render_template_string(HTML_TEMPLATE, img_path="/image")

@app.route('/download')
def download_file():
    return send_file(os.path.join(RESULT_FOLDER, "KPI_Canh_Bao.xlsx"), as_attachment=True)

@app.route('/image')
def get_image():
    return send_file(os.path.join(RESULT_FOLDER, "KPI_Canh_Bao.png"), mimetype='image/png')

@app.route('/second/process', methods=['POST'])
def second_process():
    if 'file' not in request.files:
        return "Không có file"
    file = request.files['file']
    if file.filename == '':
        return "Chưa chọn file"

    file_path = os.path.join(SECOND_UPLOAD_FOLDER, file.filename)
    file.save(file_path)

    df = pd.read_excel(file_path, header=[0, 1])
    df.columns = ['_'.join([str(i).strip() for i in col if str(i) != 'nan']) for col in df.columns.values]
    df = df.copy()

    expected_cols = [
        'Tuyến_Tuyến',
        'Phát thành công 501_Ngày',
        'Phát thành công 501_Lũy kế tháng',
        'Tỷ lệ phát thành công đúng giờ(%)_Ngày',
        'Tỷ lệ phát thành công đúng giờ(%)_Lũy kế tháng'
    ]
    for col in expected_cols:
        if col not in df.columns:
            return f"Lỗi: Không tìm thấy cột '{col}' trong file. Các cột hiện tại: {df.columns.tolist()}"

    df = df.rename(columns={
        "Tuyến_Tuyến": "Tuyến bưu tá",
        "Phát thành công 501_Ngày": "PTC ngày",
        "Phát thành công 501_Lũy kế tháng": "PTC lũy kế tháng",
        "Tỷ lệ phát thành công đúng giờ(%)_Ngày": "Tỷ lệ đúng giờ ngày",
        "Tỷ lệ phát thành công đúng giờ(%)_Lũy kế tháng": "Tỷ lệ đúng giờ tháng"
    })

    df["PTC ngày"] = pd.to_numeric(df["PTC ngày"], errors="coerce")
    df["PTC lũy kế tháng"] = pd.to_numeric(df["PTC lũy kế tháng"], errors="coerce")
    df["Tỷ lệ đúng giờ ngày"] = pd.to_numeric(df["Tỷ lệ đúng giờ ngày"], errors="coerce")
    df["Tỷ lệ đúng giờ tháng"] = pd.to_numeric(df["Tỷ lệ đúng giờ tháng"], errors="coerce")

    df_result = df[(df["PTC lũy kế tháng"].notna()) &
                    (df["Tỷ lệ đúng giờ tháng"].notna()) &
                    (df["PTC ngày"] >= 15) &
                    (df["Tỷ lệ đúng giờ ngày"] < 100)].sort_values(by="Tỷ lệ đúng giờ ngày", ascending=False)[[
        "Tuyến bưu tá",
        "PTC ngày",
        "PTC lũy kế tháng",
        "Tỷ lệ đúng giờ ngày",
        "Tỷ lệ đúng giờ tháng"
    ]].reset_index(drop=True)

    output_path = os.path.join(SECOND_RESULT_FOLDER, "Canh_Bao_Thu_Hai.xlsx")
    df_result.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for row in range(2, ws.max_row + 1):
        values = [ws[f"B{row}"].value, ws[f"C{row}"].value, ws[f"D{row}"].value, ws[f"E{row}"].value]
        ty_le_ngay = values[2]
        ty_le_thang = values[3]
        if (values[1] is not None and values[1] < 50) or (values[3] is not None and values[3] < 85) or (values[2] is not None and values[2] < 85):
            if ty_le_ngay < 85:
                ws[f"A{row}"].fill = red_fill
                ws[f"D{row}"].fill = red_fill
            if ty_le_thang < 85:
                ws[f"A{row}"].fill = red_fill
                ws[f"E{row}"].fill = red_fill
    wb.save(output_path)

    img_path = os.path.join(SECOND_RESULT_FOLDER, "Canh_Bao_Thu_Hai.png")
    height = max(5, len(df_result) * 0.4)
    fig, ax = plt.subplots(figsize=(12, height))
    ax.axis('off')
    df_result["Tỷ lệ đúng giờ ngày"] = df_result["Tỷ lệ đúng giờ ngày"].apply(lambda x: f"{x:.2f}%")
    df_result["Tỷ lệ đúng giờ tháng"] = df_result["Tỷ lệ đúng giờ tháng"].apply(lambda x: f"{x:.2f}%")
    table_data = df_result.values.tolist()
    table = ax.table(cellText=table_data, colLabels=df_result.columns, cellLoc='center', loc='center')
    table.auto_set_font_size(False)
    table.set_fontsize(9)
    table.auto_set_column_width([0, 1, 2, 3, 4])

    for key, cell in table.get_celld().items():
        if key[0] == 0:
            cell.set_text_props(weight='bold', color='white')
            cell.set_facecolor('#4C72B0')
        elif key[0] > 0:
            row_idx = key[0] - 1
            so_phat_thang = df_result.iloc[row_idx, 2]
            ty_le_thang = float(df_result.iloc[row_idx, 4].replace('%', ''))
            ty_le_ngay = float(df_result.iloc[row_idx, 3].replace('%', ''))
            if ty_le_ngay < 85 or ty_le_thang < 85:
                if ty_le_ngay < 85 and key[1] in [0, 3]:
                    cell.set_facecolor("red")
                if ty_le_thang < 85 and key[1] in [0, 4]:
                    cell.set_facecolor("red")

    plt.savefig(img_path, bbox_inches='tight', dpi=300)

    return render_template_string(SECOND_TEMPLATE, img_path="/second/image")

@app.route('/second/download')
def second_download():
    return send_file(os.path.join(SECOND_RESULT_FOLDER, "Canh_Bao_Thu_Hai.xlsx"), as_attachment=True)

@app.route('/second/image')
def second_image():
    return send_file(os.path.join(SECOND_RESULT_FOLDER, "Canh_Bao_Thu_Hai.png"), mimetype='image/png')

@app.route('/', methods=['GET'])
def home():
    return render_template_string(HTML_TEMPLATE)

@app.route('/second', methods=['GET'])
def second_form():
    return render_template_string(SECOND_TEMPLATE)

if __name__ == '__main__':
    app.run(debug=True)
